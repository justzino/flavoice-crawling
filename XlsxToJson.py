import datetime
import json
import os
from collections import OrderedDict

from dotenv import load_dotenv
from openpyxl import load_workbook

env_path = '.env'
load_dotenv(dotenv_path=env_path, verbose=True)
env = os.getenv


class Song:

    def __init__(self, title, max_pitch, explanation, singer, genre):
        self.title = title
        self.max_pitch = max_pitch.upper()
        self.explanation = explanation
        self.singer = singer
        self.genre = genre

    def convertToDict(self):
        return OrderedDict(
            title=self.title,
            max_pitch=self.max_pitch,
            explanation=self.explanation,
            singer=[self.singer],
            genre=[self.genre]
        )


class Singer:

    def __init__(self, name, date_of_debut):
        self.name = name
        self.date_of_debut = self._organizeDateOfDebut(date_of_debut)

    def _organizeDateOfDebut(self, date):
        if not date:
            date = None
        else:
            date = str(date)
            date_list = list(map(int, date.split('.')))
            if len(date_list) == 1:
                date = str(datetime.date(date_list[0], 1, 1))
            if len(date_list) == 2:
                date = str(datetime.date(date_list[0], date_list[1], 1))
            if len(date_list) == 3:
                date = str(datetime.date(date_list[0], date_list[1], date_list[2]))
        return date

    def convertToDict(self):
        return OrderedDict(
            name=self.name,
            date_of_debut=self.date_of_debut
        )


class Genre:

    def __init__(self, name):
        self.name = name

    def convertToDict(self):
        return OrderedDict(
            name=self.name
        )


# main
# 데이터 개수 입력
n_data = 50

load_path = env('xlsx_path')  # .xlsx 파일 위치
save_path = env('save_path')

sheet_name = "2021-10-30"  # 시트 이름
save_path = save_path + f"{sheet_name}.json"  # .json 저장할 파일 위치
# Excel 처리 선언
workbook = load_workbook(load_path, data_only=True)
worksheet = workbook[sheet_name]  # 으로 불러오기

data_list = []  # 데이터 목록
for i in range(2, n_data + 1):
    title = worksheet.cell(i, 1).value
    max_pitch = worksheet.cell(i, 2).value
    explanation = worksheet.cell(i, 3).value

    singer_name = worksheet.cell(i, 4).value
    singer_date_of_debut = worksheet.cell(i, 5).value
    genre_name = worksheet.cell(i, 6).value

    # singer nested 처리
    singer = Singer(singer_name, singer_date_of_debut).convertToDict()

    # genre nested 처리
    genre = Genre(genre_name).convertToDict()
    song = Song(
        title=title,
        max_pitch=max_pitch,
        explanation=explanation,
        singer=singer,
        genre=genre
    ).convertToDict()

    data_list.append(song)

json_data = json.dumps(data_list, ensure_ascii=False)  # Json 변환: 한글 저장을 위해 ensure_ascii=False

# 저장
with open(save_path, 'w+', encoding="utf-8") as file:
    file.write(json_data)
