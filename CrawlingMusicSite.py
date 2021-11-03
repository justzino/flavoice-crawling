import os
import re
import time

import chromedriver_autoinstaller
from dotenv import load_dotenv
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException

env_path = '.env'
load_dotenv(dotenv_path=env_path, verbose=True)
env = os.getenv

mode = 0

# webdirver 설정(Chrome, Firefox 등)
chromedriver_autoinstaller.install()

driver = webdriver.Chrome()
driver.set_window_size(800, 800)

version = 0
sleepTime = 0.3
resetServerCounter = 5  # 몇번 접근하여 서버 시간 불러올지 횟수 정함
serverCounter = resetServerCounter
currentTime = None


def search_singer(singer):
    # 가수 검색하기
    driver.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/fieldset/input[1]').clear()
    driver.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/fieldset/input[1]').send_keys(singer)
    driver.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/fieldset/button[2]').click()


def check_exists_by_xpath(xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True


def find_singer_info(loadPath):
    workbook = load_workbook(loadPath, data_only=True)
    worksheet = workbook['Sheet1']  # 시트 이름으로 불러오기

    # 셀 좌표로 singer 값 불러오기
    singer = worksheet.cell(2, 4).value

    DEBUT_PATTERN = re.compile('[0-9.]*[0-9]+')
    # 첫 검색
    driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/div[1]/fieldset/input[1]').send_keys(singer)
    driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/div[1]/fieldset/button[2]').click()

    # 검색 결과 데뷔일 가져오기
    # 데뷔일 가져오기
    debut_path = ''  # 결과
    debut_path1 = '/html/body/div/div[3]/div/div[1]/div[3]/div/div[1]/dl/dd[5]/span'
    debut_path2 = '/html/body/div/div[3]/div/div[1]/div[3]/div[1]/div[1]/dl/dd[5]/span'

    # debut_path 존재 체크
    if check_exists_by_xpath(debut_path1):
        debut_path = debut_path1
    elif check_exists_by_xpath(debut_path2):
        debut_path = debut_path2
    else:
        print(f"### 에러(경로 없음) singer: {singer}")

    # debut_path 값 가져오기 -> 문자열 처리
    debut_text = driver.find_element_by_xpath(debut_path).text
    debut = DEBUT_PATTERN.findall(debut_text)[0]

    print(f"2. singer: {singer}, debut: {debut}")
    worksheet.cell(2, 5, debut)

    # 나머지 singer 정보 가져오기
    for i in range(3, 565):
        # block 방지 100개마다 휴식, 저장
        if i % 100 == 0:
            workbook.save(loadPath)
            time.sleep(10)
            workbook = load_workbook(loadPath, data_only=True)
            worksheet = workbook['Sheet1']  # 시트 이름으로 불러오기
            time.sleep(10)
        last_singer = singer

        singer = worksheet.cell(i, 4).value  # 셀 좌표로 singer 값 불러오기
        if singer == last_singer:
            worksheet.cell(i, 5, debut)
            continue

        # 가수 검색하기
        search_singer(singer)

        # debut_path 존재 체크
        if check_exists_by_xpath(debut_path1):
            debut_path = debut_path1
        elif check_exists_by_xpath(debut_path2):
            debut_path = debut_path2
        else:
            print(f"### 에러(경로 없음) singer: {singer}")
            debut = ''
            continue

        # debut_path 값 가져오기 -> 문자열 처리
        debut_text = driver.find_element_by_xpath(debut_path).text
        debut = DEBUT_PATTERN.findall(debut_text)[0]

        print(f"{i}. singer: {singer}, debut: {debut}")
        worksheet.cell(i, 5, debut)
        time.sleep(1)

    workbook.save(loadPath)

    time.sleep(1)


# main
site = env('site1')
print('페이지 키는 중..')
driver.get(site)

# Excel 처리 선언
loadPath = env('xlsx_path')

find_singer_info(loadPath)

print("크롤링 종료")
